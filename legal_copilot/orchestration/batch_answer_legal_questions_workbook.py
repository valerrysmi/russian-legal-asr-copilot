"""Batch-generate answers for legal questions stored across XLSX sheets."""

from __future__ import annotations

import argparse
import csv
import sys
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

from legal_copilot.agents.context_manager import StreamingContextManager
from legal_copilot.orchestration.answer_formatting import build_short_answer
from legal_copilot.orchestration.graph import run_legal_copilot_turn

DEFAULT_INPUT = Path("Juridicheskie_voprosy_tablica.xlsx")
DEFAULT_OUTPUT = Path("Juridicheskie_voprosy_tablica - результаты.csv")
SKIP_SHEETS = {"Оглавление"}


@dataclass
class WorkbookQuestionRow:
    sheet_name: str
    row_index: int
    number: str
    question_header: str
    answer_header: str
    articles_header: str
    question: str
    reference_answer: str
    reference_articles: str


def parse_workbook_questions(workbook_path: Path) -> list[WorkbookQuestionRow]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    rows: list[WorkbookQuestionRow] = []

    for sheet_name in workbook.sheetnames:
        if sheet_name in SKIP_SHEETS:
            continue

        worksheet = workbook[sheet_name]
        parsed_rows = _parse_question_sheet(worksheet, sheet_name)
        rows.extend(parsed_rows)

    workbook.close()
    return rows


def _parse_question_sheet(worksheet, sheet_name: str) -> list[WorkbookQuestionRow]:
    extracted_rows = list(worksheet.iter_rows(values_only=True))
    header_index = _find_header_row_index(extracted_rows)
    if header_index is None:
        return []

    header_row = [normalize_cell(cell) for cell in extracted_rows[header_index]]
    question_col = _find_column_index(header_row, ("вопрос",))
    answer_col = _find_column_index(header_row, ("ответ",))
    articles_col = _find_column_index(header_row, ("смотреть", "норм", "нпа", "стать"))
    number_col = _find_number_column_index(header_row)

    if question_col is None or answer_col is None or articles_col is None:
        return []

    question_header = header_row[question_col]
    answer_header = header_row[answer_col]
    articles_header = header_row[articles_col]

    rows: list[WorkbookQuestionRow] = []
    for row_index, raw_row in enumerate(extracted_rows[header_index + 1 :], start=header_index + 2):
        normalized_row = [normalize_cell(cell) for cell in raw_row]
        question = normalized_row[question_col] if question_col < len(normalized_row) else ""
        reference_answer = normalized_row[answer_col] if answer_col < len(normalized_row) else ""
        reference_articles = normalized_row[articles_col] if articles_col < len(normalized_row) else ""
        number = normalized_row[number_col] if number_col is not None and number_col < len(normalized_row) else ""

        if not question and not reference_answer and not reference_articles:
            continue
        if not question:
            continue
        if _looks_like_section_marker(question, reference_answer, reference_articles):
            continue
        if _looks_like_repeated_header(question, reference_answer, reference_articles, number):
            continue

        rows.append(
            WorkbookQuestionRow(
                sheet_name=sheet_name,
                row_index=row_index,
                number=number,
                question_header=question_header,
                answer_header=answer_header,
                articles_header=articles_header,
                question=question,
                reference_answer=reference_answer,
                reference_articles=reference_articles,
            )
        )

    return rows


def normalize_cell(value) -> str:
    if value is None:
        return ""
    return " ".join(str(value).split()).strip()


def _find_header_row_index(rows: list[tuple]) -> int | None:
    for index, row in enumerate(rows):
        normalized = [normalize_cell(cell).lower() for cell in row]
        row_text = " | ".join(normalized)
        if "вопрос" in row_text and "ответ" in row_text:
            return index
    return None


def _find_column_index(header_row: list[str], markers: tuple[str, ...]) -> int | None:
    lowered = [cell.lower() for cell in header_row]
    for index, cell in enumerate(lowered):
        if any(marker in cell for marker in markers):
            return index
    return None


def _find_number_column_index(header_row: list[str]) -> int | None:
    lowered = [cell.lower() for cell in header_row]
    for index, cell in enumerate(lowered):
        if cell in {"№", "no", "n"}:
            return index
    return None


def _looks_like_section_marker(question: str, reference_answer: str, reference_articles: str) -> bool:
    lowered = question.lower().strip()
    if lowered.startswith("вопрос (") and not reference_answer and not reference_articles:
        return True
    if lowered in {"вопрос", "вопросы от обычного человека", "вопросы, требующие обращения к нескольким правовым документам"}:
        return True
    return False


def _looks_like_repeated_header(question: str, reference_answer: str, reference_articles: str, number: str) -> bool:
    lowered_question = question.lower().strip()
    lowered_answer = reference_answer.lower().strip()
    lowered_articles = reference_articles.lower().strip()
    lowered_number = number.lower().strip()

    if lowered_number in {"№", "n", "no"}:
        return True
    if "вопрос" in lowered_question and "ответ" in lowered_answer:
        return True
    if lowered_question.startswith("вопрос") and ("норм" in lowered_articles or "где смотреть" in lowered_articles):
        return True
    return False


def collect_used_articles(turn_result) -> str:
    if not turn_result.retrieved_context:
        return ""

    used_articles: list[str] = []
    for hit in turn_result.retrieved_context.result.hits[:5]:
        used_articles.append(f"ст. {hit.article_number} ГК РФ - {hit.title}")
    return "; ".join(used_articles)


def generate_batch_answers_from_workbook(
    input_path: Path,
    output_path: Path,
    *,
    limit: int | None = None,
) -> int:
    rows = parse_workbook_questions(input_path)
    if limit is not None:
        rows = rows[:limit]

    if not rows:
        print(f"[batch-xlsx] no question rows found in {input_path}")
        return 0

    print(f"[batch-xlsx] loaded {len(rows)} questions from {input_path}")
    output_path.parent.mkdir(parents=True, exist_ok=True)

    with output_path.open("w", encoding="utf-8-sig", newline="") as csv_file:
        writer = csv.DictWriter(
            csv_file,
            fieldnames=[
                "sheet_name",
                "row_index",
                "number",
                "question_header",
                "answer_header",
                "articles_header",
                "question",
                "reference_answer",
                "reference_articles",
                "legal_domain",
                "gk_coverage",
                "can_answer_from_civil_code",
                "short_generated_answer",
                "generated_answer",
                "used_articles",
                "route",
                "answer_source",
                "answer_generation_error",
            ],
        )
        writer.writeheader()

        for index, row in enumerate(rows, start=1):
            print(f"[batch-xlsx] ({index}/{len(rows)}) [{row.sheet_name}] {row.question}")
            session = StreamingContextManager(session_id=f"batch-xlsx-{index}")
            result = run_legal_copilot_turn(
                row.question,
                context_manager=session,
                session_id=f"batch-xlsx-{index}",
            )
            legal_domain_assessment = result.legal_domain_assessment
            generated_answer = (result.answer_text or "").strip()
            short_generated_answer = build_short_answer(generated_answer)
            used_articles = collect_used_articles(result)

            writer.writerow(
                {
                    "sheet_name": row.sheet_name,
                    "row_index": row.row_index,
                    "number": row.number,
                    "question_header": row.question_header,
                    "answer_header": row.answer_header,
                    "articles_header": row.articles_header,
                    "question": row.question,
                    "reference_answer": row.reference_answer,
                    "reference_articles": row.reference_articles,
                    "legal_domain": legal_domain_assessment.primary_domain_label if legal_domain_assessment else "",
                    "gk_coverage": legal_domain_assessment.gk_coverage if legal_domain_assessment else "",
                    "can_answer_from_civil_code": (
                        str(legal_domain_assessment.can_answer_from_civil_code) if legal_domain_assessment else ""
                    ),
                    "short_generated_answer": short_generated_answer,
                    "generated_answer": generated_answer,
                    "used_articles": used_articles,
                    "route": result.route,
                    "answer_source": result.answer_source or "",
                    "answer_generation_error": result.answer_generation_error or "",
                }
            )

    print(f"[batch-xlsx] saved results to {output_path}")
    return len(rows)


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Generate answers for all questions from a multi-sheet XLSX workbook."
    )
    parser.add_argument(
        "--input",
        type=Path,
        default=DEFAULT_INPUT,
        help="Path to the source XLSX workbook with questions.",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=DEFAULT_OUTPUT,
        help="Path to the output CSV file.",
    )
    parser.add_argument(
        "--limit",
        type=int,
        help="Optional limit for the number of processed questions.",
    )
    return parser


def main() -> int:
    args = build_parser().parse_args()
    input_path = args.input.resolve()
    output_path = args.output.resolve()

    if not input_path.exists():
        print(f"[batch-xlsx] input file not found: {input_path}", file=sys.stderr)
        return 1

    generate_batch_answers_from_workbook(input_path, output_path, limit=args.limit)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
