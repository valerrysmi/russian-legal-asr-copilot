"""Format batch-generated legal answers into a readable XLSX report."""

from __future__ import annotations

import argparse
import csv
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from legal_copilot.orchestration.answer_formatting import build_short_answer, normalize_text

DEFAULT_INPUT = Path("Юридические вопросы - результаты.csv")
DEFAULT_OUTPUT = Path("Юридические вопросы - результаты.xlsx")

COLUMNS = [
    ("sheet_name", "Лист / источник", 24),
    ("legal_domain", "Отрасль права", 20),
    ("gk_coverage", "Покрытие ГК РФ", 16),
    ("can_answer_from_civil_code", "Можно ответить только по ГК", 16),
    ("question", "Вопрос", 44),
    ("reference_answer", "Референсный ответ", 40),
    ("reference_articles", "Референсные статьи", 24),
    ("short_generated_answer", "Краткий ответ модели", 36),
    ("generated_answer", "Ответ модели", 44),
    ("used_articles", "Статьи модели", 44),
]


def load_rows(csv_path: Path) -> list[dict[str, str]]:
    with csv_path.open("r", encoding="utf-8-sig", newline="") as csv_file:
        return list(csv.DictReader(csv_file))


def write_xlsx(rows: list[dict[str, str]], output_path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Ответы"
    worksheet.freeze_panes = "A2"

    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    header_font = Font(bold=True)
    wrap_alignment = Alignment(wrap_text=True, vertical="top")

    for column_index, (_key, title, width) in enumerate(COLUMNS, start=1):
        cell = worksheet.cell(row=1, column=column_index, value=title)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = wrap_alignment
        worksheet.column_dimensions[get_column_letter(column_index)].width = width

    for row_index, row in enumerate(rows, start=2):
        normalized_row = dict(row)
        if not normalized_row.get("short_generated_answer"):
            normalized_row["short_generated_answer"] = build_short_answer(normalized_row.get("generated_answer"))

        for column_index, (key, _title, _width) in enumerate(COLUMNS, start=1):
            cell = worksheet.cell(row=row_index, column=column_index, value=normalize_text(normalized_row.get(key)))
            cell.alignment = wrap_alignment

    worksheet.auto_filter.ref = worksheet.dimensions
    workbook.save(output_path)


def format_batch_answers_xlsx(input_path: Path, output_path: Path) -> int:
    rows = load_rows(input_path)
    if not rows:
        print(f"[format-xlsx] no rows found in {input_path}")
        return 0

    output_path.parent.mkdir(parents=True, exist_ok=True)
    write_xlsx(rows, output_path)
    print(f"[format-xlsx] saved xlsx report to {output_path}")
    return len(rows)


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Convert batch answer CSV into a readable XLSX report."
    )
    parser.add_argument(
        "--input",
        type=Path,
        default=DEFAULT_INPUT,
        help="Path to the batch CSV file.",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=DEFAULT_OUTPUT,
        help="Path to the XLSX report.",
    )
    return parser


def main() -> int:
    args = build_parser().parse_args()
    input_path = args.input.resolve()
    output_path = args.output.resolve()

    if not input_path.exists():
        print(f"[format-xlsx] input file not found: {input_path}", file=sys.stderr)
        return 1

    format_batch_answers_xlsx(input_path, output_path)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
